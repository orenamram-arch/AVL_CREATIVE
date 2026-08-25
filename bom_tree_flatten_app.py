"""
bom_tree_flatten_app.py
========================
Streamlit app: uploads a multi-level BOM tree export (LEVEL / ID / NEXT_ASSY
pattern) and produces a flat workbook where each part appears once, with its
QPA (Quantity Per Assembly) broken out into one column per parent assembly.

Deploy:
  1. Push this file + requirements.txt to a GitHub repo.
  2. On https://share.streamlit.io -> New app -> pick the repo/branch ->
     main file path: bom_tree_flatten_app.py -> Deploy.

Run locally:
  pip install -r requirements.txt
  streamlit run bom_tree_flatten_app.py
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BOM Tree Flattener", page_icon="\U0001F5C2\uFE0F", layout="wide")

INFO_COLS_PRIORITY = [
    "ID", "REV", "CN", "DESCRIPTION", "UOM", "TYPE", "STATUS",
    "OWNER", "GROUP", "BUY_STATUS", "PRD_TYPE", "CATEGORY",
]
REQUIRED_COLS = {"LEVEL", "ID", "NEXT_ASSY"}

# Different ERP exports spell the same column differently. Map known
# variants to the canonical name the rest of the app expects.
HEADER_ALIASES = {
    "DESC": "DESCRIPTION",
    "DESC.": "DESCRIPTION",
    "DESCR": "DESCRIPTION",
    "PART_DESCRIPTION": "DESCRIPTION",
    "PARENT": "NEXT_ASSY",
    "PARENT_ID": "NEXT_ASSY",
    "NEXTASSY": "NEXT_ASSY",
    "NEXT_ASSEMBLY": "NEXT_ASSY",
    "PART_NUMBER": "ID",
    "PART_NO": "ID",
    "PN": "ID",
    "ITEM_ID": "ID",
    "QTY": "ORIG_QTY",
    "QUANTITY": "ORIG_QTY",
}


def canonicalize_header(h) -> str:
    """Normalize a raw header cell to the canonical column name we expect."""
    raw = str(h).strip()
    stripped = raw.rstrip(".").strip()
    key = re.sub(r"\s+", "_", stripped.upper())
    return HEADER_ALIASES.get(key, key)


# ----------------------------------------------------------------------
# Core logic (same engine as the standalone script)
# ----------------------------------------------------------------------

def sanitize_sheet_name(name: str, used_names: set) -> str:
    name = str(name)
    name = re.sub(r'[:\\/?*\[\]]', "_", name)
    name = name.strip() or "Sheet"
    name = name[:31]
    base = name
    n = 1
    while name in used_names:
        suffix = f"_{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used_names.add(name)
    return name


def load_sheet_as_table(raw: pd.DataFrame) -> pd.DataFrame:
    header = [canonicalize_header(h) for h in raw.iloc[0].tolist()]
    data = raw.iloc[1:].copy()
    data.columns = header
    data = data.reset_index(drop=True)
    data = data.dropna(how="all")
    # If a canonicalization collision produced duplicate column names, keep
    # the first occurrence of each so downstream selection stays unambiguous.
    data = data.loc[:, ~data.columns.duplicated()]
    return data


def parse_tree_nodes(block: pd.DataFrame, qty_col: str):
    """
    Parse the raw indented BOM export into an actual tree of row-nodes.

    Uses LEVEL for depth and NEXT_ASSY for the authoritative parent ID —
    NOT pure row-adjacency alone. Some exports interleave an alternate
    ("ALT") sibling row between a parent and that parent's own children
    (the parent's real children appear later, after the alternate), which
    breaks simple stack-based indent parsing. Matching each child to the
    nearest PRECEDING row whose ID equals the child's own NEXT_ASSY value
    (at the expected parent level) handles that correctly, while still
    keeping each occurrence of a repeated assembly ID as its own distinct
    node (so multi-use assemblies split properly rather than merge).
    """
    block = block.reset_index(drop=True)
    has_next_assy = "NEXT_ASSY" in block.columns
    nodes = []
    for i in range(len(block)):
        row = block.iloc[i]
        lvl_raw = pd.to_numeric(row.get("LEVEL"), errors="coerce")
        if pd.isna(lvl_raw):
            continue
        lvl = int(lvl_raw)
        qty_val = pd.to_numeric(row.get(qty_col), errors="coerce") if qty_col in row.index else None
        seq_val = pd.to_numeric(row.get("SEQ"), errors="coerce") if "SEQ" in row.index else None
        next_assy_val = None
        if has_next_assy and pd.notna(row.get("NEXT_ASSY")):
            next_assy_val = str(row.get("NEXT_ASSY")).strip()

        parent_idx = None
        if lvl > 0:
            target_level = lvl - 1
            if next_assy_val:
                for j in range(len(nodes) - 1, -1, -1):
                    if nodes[j]["level"] == target_level and nodes[j]["id"] == next_assy_val:
                        parent_idx = nodes[j]["idx"]
                        break
            if parent_idx is None:
                # No NEXT_ASSY match (missing value, or a data quirk) — fall
                # back to the nearest preceding row at the expected level.
                for j in range(len(nodes) - 1, -1, -1):
                    if nodes[j]["level"] == target_level:
                        parent_idx = nodes[j]["idx"]
                        break

        nodes.append(
            {
                "idx": len(nodes),
                "id": str(row.get("ID")),
                "level": lvl,
                "parent_idx": parent_idx,
                "qty": qty_val,
                "seq": seq_val,
                "row": row,
            }
        )
    return nodes


def flatten_single_tree(block: pd.DataFrame, qty_col: str, collapsed_ids=None):
    """
    collapsed_ids: optional set of assembly ID strings the user has decided
    to purchase complete as one unit ("buy as assembly"). Every occurrence
    of such an assembly loses its own children — both as rows (unless that
    child ID is also used somewhere outside any collapsed subtree) and as
    its own QPA_in_<id> column if it happens to itself be an assembly — but
    the collapsed assembly's own row, and its own QPA value under ITS
    parent(s), are untouched: it simply becomes a plain end-item line like
    any other part. Because pruning walks actual tree NODES (not IDs), a
    child shared with a different, non-collapsed assembly is never
    affected — that other assembly is a separate node entirely.
    """
    if qty_col not in block.columns:
        raise ValueError(f"Quantity column '{qty_col}' not found in source data.")

    nodes = parse_tree_nodes(block, qty_col)
    if not nodes:
        raise ValueError("No valid LEVEL/ID rows found in this tree block.")

    children_of_full = {}
    for n in nodes:
        if n["parent_idx"] is not None:
            children_of_full.setdefault(n["parent_idx"], []).append(n["idx"])
    for parent_idx, kids in children_of_full.items():
        children_of_full[parent_idx] = sorted(
            kids, key=lambda idx: (nodes[idx]["seq"] if pd.notna(nodes[idx]["seq"]) else 0, nodes[idx]["id"])
        )

    # Labels are always computed from the FULL (unpruned) tree, so a label
    # never changes meaning depending on which decisions happen to be
    # active — the legend and the flat table always agree on names.
    assembly_idxs_full = list(children_of_full.keys())
    id_occurrence_count = {}
    for idx in assembly_idxs_full:
        aid = nodes[idx]["id"]
        id_occurrence_count[aid] = id_occurrence_count.get(aid, 0) + 1

    labels, used_labels = {}, set()
    for idx in assembly_idxs_full:
        n = nodes[idx]
        if id_occurrence_count[n["id"]] == 1:
            label = f"QPA_in_{n['id']} (L{n['level']})"
        else:
            parent_id = nodes[n["parent_idx"]]["id"] if n["parent_idx"] is not None else "ROOT"
            label = f"QPA_in_{n['id']} [under {parent_id}] (L{n['level']})"
        final = label
        i = 2
        while final in used_labels:
            final = f"{label} #{i}"
            i += 1
        used_labels.add(final)
        labels[idx] = final

    # Full-tree DFS order (root first, each child immediately followed by
    # its own children) — used for the legend, which always shows the
    # complete structure so the user can toggle decisions either way.
    order_idx_full = []

    def dfs_full(node_idx):
        if node_idx in children_of_full:
            order_idx_full.append(node_idx)
        for c in children_of_full.get(node_idx, []):
            dfs_full(c)

    dfs_full(0)

    # Apply purchase decisions: find every occurrence-node of a collapsed
    # ID, then walk its descendants (via the FULL tree) and mark them
    # pruned. A collapsed node's own children_of entry is dropped so it no
    # longer produces a column; pruned descendants lose both their column
    # (if they had one) and, unless reachable via some other un-pruned
    # path, their row.
    collapsed_ids = {str(c) for c in collapsed_ids} if collapsed_ids else set()
    collapsed_idxs = {idx for idx in assembly_idxs_full if nodes[idx]["id"] in collapsed_ids}

    pruned_idxs = set()

    def collect_descendants(idx):
        for c in children_of_full.get(idx, []):
            if c not in pruned_idxs:
                pruned_idxs.add(c)
                collect_descendants(c)

    for idx in collapsed_idxs:
        collect_descendants(idx)

    children_of_active = {
        k: v for k, v in children_of_full.items() if k not in collapsed_idxs and k not in pruned_idxs
    }
    assembly_idxs_active = list(children_of_active.keys())

    order_idx_active = []

    def dfs_active(node_idx):
        if node_idx in children_of_active:
            order_idx_active.append(node_idx)
        for c in children_of_active.get(node_idx, []):
            dfs_active(c)

    dfs_active(0)

    child_qty = {}
    for idx in assembly_idxs_active:
        d = {}
        for c_idx in children_of_active[idx]:
            c = nodes[c_idx]
            q = c["qty"] if c["qty"] is not None and pd.notna(c["qty"]) else 0
            d[c["id"]] = d.get(c["id"], 0) + q
        child_qty[idx] = d

    info_cols = [c for c in INFO_COLS_PRIORITY if c in block.columns]
    part_first_row = {}
    for n in nodes:
        if n["level"] > 0 and n["idx"] not in pruned_idxs and n["id"] not in part_first_row:
            part_first_row[n["id"]] = n["row"]

    all_part_ids = sorted(part_first_row.keys())
    data_rows = []
    for pid in all_part_ids:
        row = part_first_row[pid]
        rec = {c: row.get(c) for c in info_cols}
        rec["ID"] = pid
        rec["BUY_AS_ASSEMBLY"] = "Y" if pid in collapsed_ids else ""
        data_rows.append(rec)
    flat = pd.DataFrame(data_rows)

    usage_count = [sum(1 for idx in order_idx_active if pid in child_qty[idx]) for pid in all_part_ids]
    flat["USED_IN_N_ASSEMBLIES"] = usage_count

    for idx in order_idx_active:
        d = child_qty[idx]
        flat[labels[idx]] = [d.get(pid, None) for pid in all_part_ids]

    flat = flat.sort_values("ID").reset_index(drop=True)

    legend_rows = []
    for idx in order_idx_full:
        n = nodes[idx]
        row = n["row"]
        parent_id = nodes[n["parent_idx"]]["id"] if n["parent_idx"] is not None else ""
        if idx in collapsed_idxs:
            status = "BUY AS ASSEMBLY — children hidden"
        elif idx in pruned_idxs:
            status = "inside a purchased assembly"
        else:
            status = ""
        legend_rows.append(
            {
                "LEVEL": n["level"],
                "ASSEMBLY_ID": n["id"],
                "DESCRIPTION": row.get("DESCRIPTION", ""),
                "REV": row.get("REV", ""),
                "PARENT_ID": parent_id,
                "COLUMN_LABEL": labels[idx] if idx in children_of_active else "",
                "BUY_AS_ASSEMBLY": n["id"] in collapsed_ids,
                "STATUS": status,
            }
        )
    legend = pd.DataFrame(legend_rows)

    return flat, legend


def format_workbook(wb):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    qpa_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row == 0:
            continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_val = str(cell.value or "")
            is_qpa = header_val.startswith("QPA_in_")
            max_len = len(header_val)
            for row_idx in range(2, ws.max_row + 1):
                body_cell = ws.cell(row=row_idx, column=col_idx)
                body_cell.font = body_font
                if is_qpa:
                    body_cell.alignment = Alignment(horizontal="center")
                    if body_cell.value not in (None, ""):
                        body_cell.fill = qpa_fill
                val_len = len(str(body_cell.value)) if body_cell.value is not None else 0
                max_len = max(max_len, val_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        ws.row_dimensions[1].height = 30
    return wb


def build_output_workbook(entries: list):
    """
    entries: list of {"root_id", "description", "source_sheet", "flat_df", "legend_df"}
    Builds the final formatted multi-sheet workbook bytes.
    """
    used_names = set()
    summary_rows = []

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for e in entries:
            out_sheet = sanitize_sheet_name(e["root_id"], used_names)
            e["flat_df"].to_excel(writer, sheet_name=out_sheet, index=False)

            legend_sheet = ""
            if not e["legend_df"].empty:
                legend_sheet = sanitize_sheet_name(f"{e['root_id']}_legend", used_names)
                e["legend_df"].to_excel(writer, sheet_name=legend_sheet, index=False)

            summary_rows.append(
                {
                    "ROOT_ID": e["root_id"],
                    "DESCRIPTION": e["description"],
                    "SOURCE_SHEET": e["source_sheet"],
                    "UNIQUE_PARTS": len(e["flat_df"]),
                    "OUTPUT_SHEET": out_sheet,
                    "LEGEND_SHEET": legend_sheet,
                }
            )

        summary_df = pd.DataFrame(summary_rows)
        summary_sheet = sanitize_sheet_name("Summary", used_names)
        summary_df.to_excel(writer, sheet_name=summary_sheet, index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    if "Summary" in wb.sheetnames:
        wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)
    wb = format_workbook(wb)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue(), summary_rows


@st.cache_data(show_spinner=False)
def parse_workbook_roots(file_bytes: bytes):
    """
    Splits the uploaded workbook into one raw block per LEVEL-0 root,
    without flattening yet — flattening happens reactively per root based
    on the current purchase decisions (collapsed_ids) picked in the UI.
    Returns a list of {sheet_name, root_id, description, block}.
    """
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    roots = []
    warnings = []
    for sheet_name in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        if raw.empty:
            continue
        data = load_sheet_as_table(raw)

        missing = REQUIRED_COLS - set(data.columns)
        if missing:
            warnings.append(f"Sheet '{sheet_name}' skipped — missing columns: {missing}")
            continue

        level_numeric = pd.to_numeric(data["LEVEL"], errors="coerce")
        root_idx = data.index[level_numeric == 0].tolist()
        if not root_idx:
            warnings.append(f"Sheet '{sheet_name}' skipped — no LEVEL=0 root row found")
            continue
        boundaries = root_idx + [len(data)]

        for i in range(len(root_idx)):
            start, end = boundaries[i], boundaries[i + 1]
            block = data.iloc[start:end].copy()
            root_row = block.iloc[0]
            roots.append(
                {
                    "sheet_name": sheet_name,
                    "root_id": str(root_row["ID"]),
                    "description": str(root_row.get("DESCRIPTION", "")),
                    "block": block,
                }
            )
    return roots, warnings


# ----------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------

st.title("\U0001F5C2\uFE0F BOM Tree Flattener")
st.caption(
    "Upload a multi-level BOM tree export (LEVEL / ID / NEXT_ASSY columns). "
    "Each level-0 root becomes its own output sheet, with every part appearing "
    "once and its QPA per parent assembly shown in separate columns."
)

with st.sidebar:
    st.header("Settings")
    qty_col = st.selectbox(
        "QPA source column",
        options=["ORIG_QTY", "REQ_QTY"],
        index=0,
        help=(
            "ORIG_QTY = nominal designed BOM quantity per assembly (recommended). "
            "REQ_QTY = point-in-time MRP net requirement, net of stock/orders."
        ),
    )

uploaded = st.file_uploader("Upload BOM tree file (.xls / .xlsx)", type=["xls", "xlsx"])

if uploaded is not None:
    with st.spinner("Reading BOM tree..."):
        try:
            roots, warnings = parse_workbook_roots(uploaded.getvalue())
        except Exception as e:
            st.error(f"Failed to read file: {e}")
            st.stop()

    for w in warnings:
        st.warning(w)

    if not roots:
        st.error("No valid LEVEL=0 root products were found in this file.")
        st.stop()

    st.success(f"Found {len(roots)} root product(s).")
    st.caption(
        "For any sub-assembly you plan to purchase complete, tick **BUY_AS_ASSEMBLY** "
        "in its legend below — its own children will drop out of the flattened table "
        "(unless a child part is also used elsewhere, which stays untouched)."
    )

    entries = []
    tab_labels = [f"{r['root_id']}" for r in roots]
    tabs = st.tabs(tab_labels)

    for root, tab in zip(roots, tabs):
        with tab:
            root_id = root["root_id"]
            block = root["block"]

            try:
                base_flat, base_legend = flatten_single_tree(block, qty_col=qty_col, collapsed_ids=None)
            except ValueError as e:
                st.warning(f"Skipped: {e}")
                continue

            st.markdown(f"**{root['description']}**  ·  source sheet: `{root['sheet_name']}`")

            editor_df = base_legend[["LEVEL", "ASSEMBLY_ID", "DESCRIPTION", "REV", "PARENT_ID", "BUY_AS_ASSEMBLY"]].copy()
            edited = st.data_editor(
                editor_df,
                key=f"legend_editor_{root_id}",
                hide_index=True,
                use_container_width=True,
                disabled=["LEVEL", "ASSEMBLY_ID", "DESCRIPTION", "REV", "PARENT_ID"],
                column_config={
                    "BUY_AS_ASSEMBLY": st.column_config.CheckboxColumn(
                        "Buy as assembly?", help="Purchase this sub-assembly complete — hide its children below."
                    )
                },
            )

            collapsed_ids = set(edited.loc[edited["BUY_AS_ASSEMBLY"] == True, "ASSEMBLY_ID"].astype(str))  # noqa: E712

            try:
                flat_df, legend_df = flatten_single_tree(block, qty_col=qty_col, collapsed_ids=collapsed_ids)
            except ValueError as e:
                st.warning(f"Skipped: {e}")
                continue

            entries.append(
                {
                    "root_id": root_id,
                    "description": root["description"],
                    "source_sheet": root["sheet_name"],
                    "flat_df": flat_df,
                    "legend_df": legend_df,
                }
            )

            st.caption(f"{len(flat_df)} unique parts" + (f" · {len(collapsed_ids)} purchased as assembly" if collapsed_ids else ""))
            st.dataframe(flat_df, use_container_width=True)

    if entries:
        output_bytes, summary_rows = build_output_workbook(entries)
        st.divider()
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)
        out_name = uploaded.name.rsplit(".", 1)[0] + "_flattened.xlsx"
        st.download_button(
            "\u2b07\uFE0F Download flattened workbook",
            data=output_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload a file to get started.")
